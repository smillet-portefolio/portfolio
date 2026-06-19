#!/usr/bin/env python3
"""
update_prix_gsheet.py
Écrit les prix ET les variations % (1D / 1W / 1M / 3M / 6M / 1Y / YTD)
dans Google Sheets (onglet "Prices") avec décimales complètes.
Inclut aussi les valeurs de part des fonds de retraite Conseq (CONSEQ_GLAK/RAF/DL).
Prérequis : pip install yfinance pandas requests google-auth google-api-python-client
"""

import datetime
import os
import json
import re
import requests
import pandas as pd
import yfinance as yf

SERVICE_ACCOUNT_FILE = r"C:\Users\smill\OneDrive\Documents\banque\bourse\service_account.json"
SHEET_ID  = os.environ.get("SHEET_ID", "15w4s6chCytFKmPSpGXeYQ9fiJEVD_T5U9671Q0chn_Q")
SHEET_TAB = "Prices"
LINXEA_FILE = os.environ.get("LINXEA_FILE", r"C:\Users\smill\OneDrive\Documents\banque\bourse\linxea.xlsx")

try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    GOOGLE_OK = True
except ImportError:
    print("pip install google-auth google-api-python-client")
    GOOGLE_OK = False

# (ticker interne, nom, devise, symbole Yahoo pour l'historique des variations)
TICKERS = [
    ("ABVX.PA",  "Abivax",                        "EUR", "ABVX.PA"),
    ("FGR.PA",   "Eiffage",                       "EUR", "FGR.PA"),
    ("VIE.PA",   "Vonovia",                       "EUR", "VIE.PA"),
    ("DTE.DE",   "Deutsche Telekom",              "EUR", "DTE.DE"),
    ("EL.PA",    "EssilorLuxottica",              "EUR", "EL.PA"),
    ("RACE.MI",  "Ferrari",                       "EUR", "RACE.MI"),
    ("IDL.PA",   "ID Logistics",                  "EUR", "IDL.PA"),
    ("IFX.DE",   "Infineon",                      "EUR", "IFX.DE"),
    ("ENR.DE",   "Siemens Energy",                "EUR", "ENR.DE"),
    ("SGO.PA",   "Saint-Gobain",                  "EUR", "SGO.PA"),
    ("SU.PA",    "Schneider Electric",            "EUR", "SU.PA"),
    ("RR.L",     "Rolls-Royce",                   "GBp", "RR.L"),
    ("MSFT",     "Microsoft",                     "USD", "MSFT"),
    ("ORCL",     "Oracle",                        "USD", "ORCL"),
    ("VST",      "Vistra Corp",                   "USD", "VST"),
    ("VU.PA",    "Vusion Group",                  "EUR", "VU.PA"),
    ("SPIE.PA",  "Spie",                          "EUR", "SPIE.PA"),
    ("EMXC.DE",  "Amundi MSCI Emerging ex China", "EUR", "EMXC.DE"),
    ("LSMC.DE",  "Amundi MSCI Semiconductors",    "EUR", "LSMC.DE"),
    ("HYDE.DE",  "Invesco Hydrogen Economy",      "EUR", "HYDE.DE"),
    ("IQQH.DE",  "iShares Global Clean Energy",   "EUR", "IQQH.DE"),
    ("I500.DE",  "iShares S&P 500 Swap",          "EUR", "I500.DE"),
    ("JEDI.DE",  "VanEck Space Innovators",       "EUR", "JEDI.DE"),
    ("NUKL.DE",  "VanEck Uranium & Nuclear",      "EUR", "NUKL.DE"),
    ("XDWD.DE",  "Xtrackers MSCI World",          "EUR", "XDWD.DE"),
    ("MTPI.PA",  "iShares MSCI ex China",         "EUR", "MTPI.PA"),
    ("VGWE.DE",  "Vanguard All-World High Div",   "EUR", "VGWE.DE"),
    ("BTC/EUR",  "Bitcoin",                       "EUR", "BTC-EUR"),
    ("ETH/EUR",  "Ethereum",                      "EUR", "ETH-EUR"),
    ("EUR/CZK",  "Euro / Couronne tchèque",       "CZK", "EURCZK=X"),
    ("USD/CZK",  "Dollar / Couronne tchèque",     "CZK", "USDCZK=X"),
    ("GBP/CZK",  "Livre / Couronne tchèque",      "CZK", "GBPCZK=X"),
]

TICKERS_US = {"MSFT", "ORCL", "VST"}

# Colonnes de variations dans l'ordre voulu
VAR_KEYS = ["1D", "1W", "1M", "3M", "6M", "1Y", "YTD"]


def prix_yahoo(ticker):
    try:
        t = yf.Ticker(ticker)
        if ticker in TICKERS_US:
            info = t.info
            p = info.get("currentPrice") or info.get("regularMarketPrice")
            if p and float(p) > 0:
                return float(p)
        p = t.fast_info.last_price
        if p and p > 0:
            return float(p)
    except Exception as e:
        print(f"  [Yahoo {ticker}] {e}")
    return None


def variations_yahoo(yahoo_sym):
    """Retourne {1D,1W,1M,3M,6M,1Y,YTD} en % à partir de l'historique Yahoo."""
    vides = {k: None for k in VAR_KEYS}
    try:
        hist = yf.Ticker(yahoo_sym).history(period="2y", auto_adjust=False)
        if hist.empty:
            return vides
        closes = hist["Close"].dropna()
        if closes.empty:
            return vides

        last = float(closes.iloc[-1])
        ld = closes.index[-1]

        def close_on_or_before(target):
            sub = closes[closes.index <= target]
            return float(sub.iloc[-1]) if not sub.empty else None

        def pct(ref):
            if ref and ref > 0:
                return round((last / ref - 1) * 100, 2)
            return None

        res = dict(vides)
        res["1D"] = pct(float(closes.iloc[-2])) if len(closes) >= 2 else None
        res["1W"] = pct(close_on_or_before(ld - pd.Timedelta(days=7)))
        res["1M"] = pct(close_on_or_before(ld - pd.DateOffset(months=1)))
        res["3M"] = pct(close_on_or_before(ld - pd.DateOffset(months=3)))
        res["6M"] = pct(close_on_or_before(ld - pd.DateOffset(months=6)))
        res["1Y"] = pct(close_on_or_before(ld - pd.DateOffset(years=1)))

        ytd_start = pd.Timestamp(year=ld.year, month=1, day=1, tz=ld.tz)
        prev = closes[closes.index < ytd_start]  # dernier cours de l'an passé
        res["YTD"] = pct(float(prev.iloc[-1])) if not prev.empty else None
        return res
    except Exception as e:
        print(f"  [Var {yahoo_sym}] {e}")
        return vides


def prix_binance(symbole):
    try:
        r = requests.get(f"https://api.binance.com/api/v3/ticker/price?symbol={symbole}", timeout=10)
        r.raise_for_status()
        return float(r.json()["price"])
    except Exception as e:
        print(f"  [Binance {symbole}] {e}")
    return None


def taux_cnb():
    taux = {}
    try:
        r = requests.get(
            "https://www.cnb.cz/en/financial-markets/foreign-exchange-market"
            "/central-bank-exchange-rate-fixing/central-bank-exchange-rate-fixing/daily.txt",
            timeout=10)
        r.raise_for_status()
        for ligne in r.text.strip().split("\n")[2:]:
            parts = ligne.split("|")
            if len(parts) < 5:
                continue
            code   = parts[3].strip()
            amount = int(parts[2].strip())
            rate   = float(parts[4].strip().replace(",", ".")) / amount
            if code in ("EUR", "USD", "GBP"):
                taux[code] = rate
    except Exception as e:
        print(f"  [CNB] {e}")
    return taux


# Fonds de retraite Conseq (DPS) — valeur de part publiée sur conseq.cz
# (ticker interne pour l'onglet Prices, nom, slug de la page)
CONSEQ_FUNDS = [
    ("CONSEQ_GLAK", "Conseq globalni akciovy", "conseq-globalni-akciovy-ucastnicky-fond"),
    ("CONSEQ_RAF",  "Conseq realitni",         "conseq-realitni-ucastnicky-fond"),
    ("CONSEQ_DL",   "Conseq dluhopisovy",      "conseq-dluhopisovy-ucastnicky-fond"),
    ("CONSEQ_BOND35", "Conseq Target Bond 2035", "conseq-target-bond-2035-ucastnicky-fond"),
]

_CONSEQ_PAT = re.compile(
    r"Aktu[aá]ln[ií]\s+hodnota\s+penzijn[ií]\s+jednotky[:\s]*"
    r"([\d\s ]+,\d+)\s*CZK\s*\(([^)]+)\)",
    re.IGNORECASE,
)


def prix_conseq(slug):
    """Retourne (valeur_part_CZK, date_str) depuis la page publique du fonds Conseq."""
    try:
        url = "https://www.conseq.cz/penze/prehled-ucastnickych-fondu/" + slug
        r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=20)
        r.raise_for_status()
        text = re.sub(r"<[^>]+>", " ", r.text)
        text = text.replace("\xa0", " ").replace("&nbsp;", " ")
        text = re.sub(r"\s+", " ", text)
        m = re.search(r"hodnota penzijn[ií] jednotky[:\s]*([\d ]+,\d+)\s*CZK\s*\(([^)]+?\d{4})\)", text, re.IGNORECASE)
        if not m:
            return None, None
        nav = float(m.group(1).replace(" ", "").replace(" ", "").replace(",", "."))
        return nav, m.group(2).strip()
    except Exception as e:
        print(f"  [Conseq {slug}] {e}")
        return None, None


def fmt_prix(p):
    """Convertit un float en string avec toutes ses décimales significatives."""
    if p is None:
        return ""
    return repr(p)


def fmt_var(v):
    """Variation en string ('' si indisponible) pour USER_ENTERED."""
    if v is None:
        return ""
    return repr(v)


def collecter_prix():
    print("\n--- Taux CNB ---")
    cnb = taux_cnb()
    for devise, val in cnb.items():
        print(f"  OK {devise}/CZK -> {val}")

    print("\n--- Crypto Binance ---")
    btc = prix_binance("BTCEUR")
    eth = prix_binance("ETHEUR")
    print(f"  {'OK' if btc else 'KO'} Bitcoin  -> {btc}")
    print(f"  {'OK' if eth else 'KO'} Ethereum -> {eth}")

    print("\n--- Actions / ETFs / Variations ---")
    maintenant = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    header = ["Ticker", "Nom", "price last closure", "currency", "Last Update"] + VAR_KEYS
    rows = [header]

    for ticker, nom, currency, ysym in TICKERS:
        if ticker == "BTC/EUR":
            prix = btc
        elif ticker == "ETH/EUR":
            prix = eth
        elif ticker == "EUR/CZK":
            prix = cnb.get("EUR")
        elif ticker == "USD/CZK":
            prix = cnb.get("USD")
        elif ticker == "GBP/CZK":
            prix = cnb.get("GBP")
        else:
            prix = prix_yahoo(ticker)

        # Variations calculées depuis l'historique Yahoo (relatif → source fiable)
        var = variations_yahoo(ysym)
        var_str = "  ".join(f"{k}:{var[k]}" for k in VAR_KEYS)
        status = "OK" if prix else "KO"
        print(f"  {status} {nom:<40s} ({ticker}) -> {prix}  |  {var_str}")

        # Prix stocké comme STRING pour éviter l'arrondi Google Sheets
        rows.append(
            [ticker, nom, fmt_prix(prix), currency, maintenant if prix else ""]
            + [fmt_var(var[k]) for k in VAR_KEYS]
        )

    print("\n--- Conseq (fonds retraite) ---")
    for tick, nom, slug in CONSEQ_FUNDS:
        nav, ddate = prix_conseq(slug)
        print(f"  {'OK' if nav else 'KO'} {nom:<40s} ({tick}) -> {nav}  ({ddate})")
        rows.append(
            [tick, nom + (f" ({ddate})" if ddate else ""), fmt_prix(nav), "CZK",
             (maintenant if nav else "")]
            + ["" for _ in VAR_KEYS]
        )

    return rows


def ecrire_google_sheets(rows):
    if not GOOGLE_OK:
        print("❌ Modules Google manquants")
        return False
    try:
        SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
        sa_env = os.environ.get("GOOGLE_SERVICE_ACCOUNT")
        if sa_env:  # exécution cloud (GitHub Actions) : identifiants dans une variable/secret
            creds = service_account.Credentials.from_service_account_info(
                json.loads(sa_env), scopes=SCOPES)
        else:       # exécution locale : fichier service_account.json
            creds = service_account.Credentials.from_service_account_file(
                SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        service = build("sheets", "v4", credentials=creds)
        sheet   = service.spreadsheets()

        meta = sheet.get(spreadsheetId=SHEET_ID).execute()
        tabs = [s["properties"]["title"] for s in meta["sheets"]]
        if SHEET_TAB not in tabs:
            sheet.batchUpdate(spreadsheetId=SHEET_ID, body={
                "requests": [{"addSheet": {"properties": {"title": SHEET_TAB}}}]
            }).execute()
            print(f"  Onglet '{SHEET_TAB}' créé")

        sheet.values().clear(spreadsheetId=SHEET_ID, range=f"{SHEET_TAB}!A:Z").execute()
        sheet.values().update(
            spreadsheetId=SHEET_ID,
            range=f"{SHEET_TAB}!A1",
            valueInputOption="USER_ENTERED",
            body={"values": rows}
        ).execute()
        print(f"\n  ✅ {len(rows)-1} lignes (prix + variations) écrites")
        return True

    except FileNotFoundError:
        print(f"\n  ❌ Fichier introuvable : {SERVICE_ACCOUNT_FILE}")
        return False
    except Exception as e:
        print(f"\n  ❌ Erreur : {e}")
        return False


# ==============================================================================
# MISE A JOUR DU FICHIER LOCAL linxea.xlsx (onglet "update python")
# Etape optionnelle : ignoree si le fichier est absent ou si openpyxl manque
# (donc sans effet en execution cloud / GitHub Actions).
# ==============================================================================
ISIN_VERS_TICKER = {
    # Actions
    "FR0012333284": "ABVX.PA",
    "FR0000130452": "FGR.PA",
    "FR0000124141": "VIE.PA",
    "DE0005557508": "DTE.DE",
    "FR0000121667": "EL.PA",
    "NL0011585146": "RACE.MI",
    "FR0010929125": "IDL.PA",
    "DE0006231004": "IFX.DE",
    "DE000ENER6Y0": "ENR.DE",
    "FR0000125007": "SGO.PA",
    "FR0000121972": "SU.PA",
    "GB00B63H8491": "RR.L",
    "US5949181045": "MSFT",
    "US68389X1054": "ORCL",
    "US92840M1027": "VST",
    "FR0010282822": "VU.PA",
    # ETFs
    "LU2009202107": "EMXC.DE",
    "LU1900066033": "LSMC.DE",
    "IE00053WDH64": "HYDE.DE",
    "IE00B1XNHC34": "IQQH.DE",
    "IE00BMTX1Y45": "I500.DE",
    "FR0012757854": "SPIE.PA",
    "IE000JJPY166": "YCSH.DE",
    "IE000YU9K6K2": "JEDI.DE",
    "IE000M7V94E1": "NUKL.DE",
    "LU0476289466": "D5BI.DE",
    "IE00BJ0KDQ92": "XDWD.DE",
    "IE00B6R52259": "IUSQ.DE",
    "IE000BI8OT95": "MWRD.MI",
    "LU1900066462": "LEER.DE",
    "IE000U58J0M1": "INRE.PA",
    "IE00BMG6Z448": "MTPI.PA",
    "IE00BK5BR626": "VGWE.DE",
}


def mise_a_jour_linxea(fichier=None):
    fichier = fichier or LINXEA_FILE
    print("\n" + "-" * 55)
    print(f"  LINXEA (fichier local) : {fichier}")
    print("-" * 55)
    if not fichier or not os.path.exists(fichier):
        print("  -> Fichier linxea.xlsx absent : etape ignoree (normal en cloud).")
        return
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  -> openpyxl non installe : etape ignoree (pip install openpyxl).")
        return
    try:
        wb = load_workbook(fichier)
    except Exception as e:
        print(f"  ERREUR lecture : {e}")
        return
    if "update python" not in wb.sheetnames:
        print(f"  ERREUR : onglet 'update python' introuvable. Onglets : {wb.sheetnames}")
        return

    ws = wb["update python"]
    cnb = taux_cnb()
    btc = prix_binance("BTCEUR")
    eth = prix_binance("ETHEUR")

    # Detection de la ligne d'en-tete
    ligne_entete = 1
    for r in range(1, 10):
        if any(ws.cell(r, c).value for c in range(1, ws.max_column + 1)):
            ligne_entete = r
            break

    entetes = {ws.cell(ligne_entete, c).value: c
               for c in range(1, ws.max_column + 1)
               if ws.cell(ligne_entete, c).value}

    col_nom    = entetes.get("Nom")
    col_isin   = entetes.get("ISIN")
    col_source = entetes.get("information source")
    col_prix   = entetes.get("price last closure")
    col_date   = entetes.get("Last Update")
    col_ticker = entetes.get("Ticker")

    if not all([col_nom, col_isin, col_source, col_prix, col_date]):
        print(f"  ERREUR : colonnes manquantes. Trouvees : {list(entetes.keys())}")
        return

    if col_ticker is None:
        ws.insert_cols(col_isin + 1)
        ws.cell(ligne_entete, col_isin + 1).value = "Ticker"
        col_ticker = col_isin + 1
        if col_source > col_isin: col_source += 1
        if col_prix   > col_isin: col_prix   += 1
        if col_date   > col_isin: col_date   += 1

    maintenant = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    nb_ok = nb_ko = 0

    for row_num in range(ligne_entete + 1, ws.max_row + 1):
        nom    = ws.cell(row_num, col_nom).value
        isin   = str(ws.cell(row_num, col_isin).value or "").strip()
        source = str(ws.cell(row_num, col_source).value or "").strip().lower()
        if not nom:
            continue

        prix = None
        ticker_affiche = None
        nom_l = nom.lower()

        if source == "binance":
            if "bitcoin" in nom_l:
                prix, ticker_affiche = btc, "BTC/EUR"
            elif "etherum" in nom_l or "ethereum" in nom_l:
                prix, ticker_affiche = eth, "ETH/EUR"
        elif source == "cnb.cz":
            if "eur" in nom_l:
                prix, ticker_affiche = cnb.get("EUR"), "EUR/CZK"
            elif "usd" in nom_l:
                prix, ticker_affiche = cnb.get("USD"), "USD/CZK"
            elif "gbp" in nom_l:
                prix, ticker_affiche = cnb.get("GBP"), "GBP/CZK"
        elif source == "boursorama":
            ticker_affiche = ISIN_VERS_TICKER.get(isin)
            if ticker_affiche:
                prix = prix_yahoo(ticker_affiche)
                print(f"  {'OK' if prix else 'KO'} {nom} ({ticker_affiche}) -> {prix}")
            else:
                print(f"  ?? {nom} — ticker inconnu (ISIN: {isin})")

        if ticker_affiche:
            ws.cell(row_num, col_ticker).value = ticker_affiche
        if prix is not None:
            ws.cell(row_num, col_prix).value = prix
            ws.cell(row_num, col_date).value = maintenant
            nb_ok += 1
        elif source in ("boursorama", "binance", "cnb.cz"):
            nb_ko += 1

    wb.save(fichier)
    print(f"\n  OK : {nb_ok} mis a jour  |  KO : {nb_ko} non trouves  ->  linxea.xlsx sauvegarde.")


if __name__ == "__main__":
    print("=" * 55)
    print("  Mise à jour prix + variations → Google Sheets")
    print("=" * 55)
    rows = collecter_prix()
    ok   = ecrire_google_sheets(rows)
    if ok:
        print("\n✅ Prix et variations écrits")
    else:
        print("\n⚠️  Vérifier service_account.json")
    # Etape locale optionnelle : mise a jour du fichier linxea.xlsx
    try:
        mise_a_jour_linxea(LINXEA_FILE)
    except Exception as e:
        print(f"  [Linxea] {e}")
    print("=" * 55)
