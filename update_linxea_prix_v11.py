#!/usr/bin/env python3
"""
update_linxea_prix.py  —  version 11 (linxea.xlsx uniquement)
-------------------------------------------------------------
Sources :
  - Yahoo Finance (yfinance) pour actions, ETFs et crypto
  - Binance API pour Bitcoin et Ethereum (EUR) — prix temps réel
  - CNB pour EUR/CZK, USD/CZK et GBP/CZK

Fichier mis à jour :
  - linxea.xlsx  (onglet "update python") — prix uniquement

Usage :
  python update_linxea_prix_v11.py                 # chemin par défaut
  python update_linxea_prix_v11.py <chemin_linxea.xlsx>

Prérequis : pip install yfinance openpyxl requests
"""

import datetime
import sys
import os
import requests
import yfinance as yf
from openpyxl import load_workbook

# ─── CHEMIN DU FICHIER ────────────────────────────────────────────────────────
DOSSIER = r"C:\Users\smill\OneDrive\Documents\banque\bourse"

if len(sys.argv) >= 2:
    FICHIER_LINXEA = sys.argv[1]
else:
    FICHIER_LINXEA = os.path.join(DOSSIER, "linxea.xlsx")

# ─── CORRESPONDANCE ISIN → TICKER YAHOO ──────────────────────────────────────
ISIN_VERS_TICKER = {
    # Actions
    "FR0012333284": "ABVX.PA",
    "FR0000130452": "FGR.PA",       # Eiffage
    "FR0000124141": "VIE.PA",       # Veolia
    "DE0005557508": "DTE.DE",
    "FR0000121667": "EL.PA",
    "NL0011585146": "RACE.MI",
    "FR0010929125": "IDL.PA",
    "DE0006231004": "IFX.DE",
    "DE000ENER6Y0": "ENR.DE",
    "FR0000125007": "SGO.PA",
    "FR0000121972": "SU.PA",
    "GB00B63H8491": "RR.L",
    "US5949181045": "MSFT",
    "US68389X1054": "ORCL",
    "US92840M1027": "VST",
    "FR0010282822": "VU.PA",
    "ES0144580Y14": "IBE.MC",       # Iberdrola
    "US92537N1081": "VRT",          # Vertiv Holdings
    # ETFs
    "LU2009202107": "EMXC.DE",
    "LU1900066033": "LSMC.DE",
    "IE00053WDH64": "HYDE.DE",
    "IE00B1XNHC34": "IQQH.DE",
    "IE00BMTX1Y45": "I500.DE",
    "FR0012757854": "SPIE.PA",
    "IE000JJPY166": "YCSH.DE",
    "IE000YU9K6K2": "JEDI.DE",
    "IE000M7V94E1": "NUKL.DE",
    "LU0476289466": "D5BI.DE",
    "IE00BJ0KDQ92": "XDWD.DE",
    "IE00B6R52259": "IUSQ.DE",
    "IE000BI8OT95": "MWRD.MI",
    "LU1900066462": "LEER.DE",
    "IE000U58J0M1": "INRE.PA",
    "IE00BMG6Z448": "MTPI.PA",
    "IE00BK5BR626": "VGWE.DE",
    "IE00BM67HK77": "XDWH.DE",       # Xtrackers MSCI World Health Care
}

# Tickers US avec prix after-hours (on enregistre toujours le prix de clôture)
TICKERS_US = {"MSFT", "ORCL", "VST", "VRT"}


# ─── RÉCUPÉRATION DES PRIX ────────────────────────────────────────────────────

def prix_yahoo(ticker):
    """Retourne le prix de clôture du marché régulier (repli multi-place)."""
    def _one(sym):
        try:
            t = yf.Ticker(sym)
            if sym in TICKERS_US:
                info = t.info
                reg = info.get("currentPrice") or info.get("regularMarketPrice")
                if reg and float(reg) > 0:
                    return round(float(reg), 4)
            p = t.fast_info.last_price
            if p and p > 0:
                return round(float(p), 4)
        except Exception as e:
            print(f"    [Yahoo {sym}] {e}")
        return None

    # Ticker déjà suffixé (.DE/.PA/-/=) : essai direct
    if any(c in ticker for c in (".", "-", "=")):
        return _one(ticker)
    # Sinon repli multi-place pour un nouveau titre européen sans suffixe
    for suf in ["", ".DE", ".MI", ".PA", ".AS", ".L", ".SW", ".F"]:
        p = _one(ticker + suf)
        if p is not None:
            if suf:
                print(f"    [resolve] {ticker} -> {ticker + suf}")
            return p
    return None


def prix_binance(symbole):
    """Retourne le prix temps réel depuis Binance."""
    try:
        r = requests.get(
            f"https://api.binance.com/api/v3/ticker/price?symbol={symbole}",
            timeout=10)
        r.raise_for_status()
        return round(float(r.json()["price"]), 2)
    except Exception as e:
        print(f"    [Binance {symbole}] {e}")
    return None


def taux_cnb():
    """Retourne les taux EUR, USD, GBP depuis la CNB."""
    taux = {}
    try:
        r = requests.get(
            "https://www.cnb.cz/en/financial-markets/foreign-exchange-market"
            "/central-bank-exchange-rate-fixing/central-bank-exchange-rate-fixing/daily.txt",
            timeout=10)
        r.raise_for_status()
        for ligne in r.text.strip().split("\n")[2:]:
            parts = ligne.split("|")
            if len(parts) < 5:
                continue
            code = parts[3].strip()
            amount = int(parts[2].strip())
            rate = float(parts[4].strip().replace(",", ".")) / amount
            if code in ("EUR", "USD", "GBP"):
                taux[code] = round(rate, 4)
    except Exception as e:
        print(f"    [CNB] {e}")
    return taux


# ─── MISE À JOUR LINXEA.XLSX (onglet "update python") ────────────────────────

def mise_a_jour_linxea(fichier, btc, eth, cnb):
    print(f"\n{'─'*55}")
    print(f"  LINXEA : {fichier}")
    print(f"{'─'*55}")
    try:
        wb = load_workbook(fichier)
    except FileNotFoundError:
        print("  ERREUR : fichier introuvable")
        return False

    if "update python" not in wb.sheetnames:
        print(f"  ERREUR : onglet 'update python' introuvable. Onglets : {wb.sheetnames}")
        return False

    ws = wb["update python"]

    # Détecter la ligne d'en-tête
    ligne_entete = 1
    for r in range(1, 10):
        if any(ws.cell(r, c).value for c in range(1, ws.max_column + 1)):
            ligne_entete = r
            break

    entetes = {ws.cell(ligne_entete, c).value: c
               for c in range(1, ws.max_column + 1)
               if ws.cell(ligne_entete, c).value}

    col_nom    = entetes.get("Nom")
    col_isin   = entetes.get("ISIN")
    col_source = entetes.get("information source")
    col_prix   = entetes.get("price last closure")
    col_date   = entetes.get("Last Update")
    col_ticker = entetes.get("Ticker")

    if not all([col_nom, col_isin, col_source, col_prix, col_date]):
        print(f"  ERREUR : colonnes manquantes. Trouvées : {list(entetes.keys())}")
        return False

    if col_ticker is None:
        ws.insert_cols(col_isin + 1)
        ws.cell(ligne_entete, col_isin + 1).value = "Ticker"
        col_ticker = col_isin + 1
        if col_source > col_isin: col_source += 1
        if col_prix   > col_isin: col_prix   += 1
        if col_date   > col_isin: col_date   += 1

    maintenant = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    nb_ok = nb_ko = 0

    for row_num in range(ligne_entete + 1, ws.max_row + 1):
        nom    = ws.cell(row_num, col_nom).value
        isin   = str(ws.cell(row_num, col_isin).value or "").strip()
        source = str(ws.cell(row_num, col_source).value or "").strip().lower()
        if not nom:
            continue

        prix = None
        ticker_affiche = None
        nom_l = nom.lower()

        if source == "binance":
            if "bitcoin" in nom_l:
                prix, ticker_affiche = btc, "BTC/EUR"
            elif "etherum" in nom_l or "ethereum" in nom_l:
                prix, ticker_affiche = eth, "ETH/EUR"

        elif source == "cnb.cz":
            if "eur" in nom_l:
                prix, ticker_affiche = cnb.get("EUR"), "EUR/CZK"
            elif "usd" in nom_l:
                prix, ticker_affiche = cnb.get("USD"), "USD/CZK"
            elif "gbp" in nom_l:
                prix, ticker_affiche = cnb.get("GBP"), "GBP/CZK"

        elif source == "boursorama":
            ticker_affiche = ISIN_VERS_TICKER.get(isin)
            if ticker_affiche:
                prix = prix_yahoo(ticker_affiche)
                print(f"  {'OK' if prix else 'KO'} {nom} ({ticker_affiche}) -> {prix}")
            else:
                print(f"  ?? {nom} — ticker inconnu (ISIN: {isin})")

        if ticker_affiche:
            ws.cell(row_num, col_ticker).value = ticker_affiche
        if prix is not None:
            ws.cell(row_num, col_prix).value = prix
            ws.cell(row_num, col_date).value = maintenant
            nb_ok += 1
        elif source in ("boursorama", "binance", "cnb.cz"):
            nb_ko += 1

    wb.save(fichier)
    print(f"\n  OK : {nb_ok} mis à jour  |  KO : {nb_ko} non trouvés  →  sauvegardé.")
    return True


# ─── POINT D'ENTRÉE ───────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 55)
    print("  Mise à jour des prix Linxea — version 11 (linxea.xlsx)")
    print("=" * 55)

    print("\n--- Taux CNB ---")
    cnb = taux_cnb()
    for devise, val in cnb.items():
        print(f"  OK {devise}/CZK -> {val}")
    if not cnb:
        print("  KO Impossible de joindre la CNB")

    print("\n--- Crypto Binance (EUR) ---")
    btc = prix_binance("BTCEUR")
    eth = prix_binance("ETHEUR")
    print(f"  {'OK' if btc else 'KO'} Bitcoin  -> {btc} EUR")
    print(f"  {'OK' if eth else 'KO'} Ethereum -> {eth} EUR")

    print("\n--- Actions / ETFs / Crypto (Yahoo Finance) ---")
    ok = mise_a_jour_linxea(FICHIER_LINXEA, btc, eth, cnb)

    print(f"\n{'='*55}")
    print("  Terminé." if ok else "  Terminé avec erreurs.")
    print(f"{'='*55}")
    sys.exit(0 if ok else 1)
