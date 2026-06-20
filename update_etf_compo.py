# -*- coding: utf-8 -*-
"""
================================================================================
 update_etf_compo.py  —  Mise a jour des compositions ETF (repartition_ETFs.xlsx)
================================================================================

Met a jour les 3 onglets de repartition_ETFs.xlsx :
   1) "by country"      -> repartition geographique (pays) de chaque ETF
   2) "by sector"       -> repartition sectorielle de chaque ETF
   3) "TOP 10 position" -> 10 premieres lignes de chaque ETF

LISTE DYNAMIQUE
---------------
La liste des ETF a mettre a jour est construite DYNAMIQUEMENT a partir des ETF
reellement detenus dans le portefeuille (meme logique que update_prix_gsheet.py) :
  - lecture des positions depuis l'onglet "Data" du Google Sheet
    (cles book2_all_rows + book2_extra_rows),
  - on ne garde que les lignes dont le ticker est un ETF (present dans ETF_META,
    lu depuis dashboard_portefeuille.html) ou dont typeInv == "ETF".
Les ETF nouvellement ajoutes sont donc pris en compte automatiquement ; les ETF
vendus (absents du portefeuille) sont ignores.

SOURCES DE DONNEES
------------------
  - justETF (par ISIN)  : source universelle et fiable pour TOUS les ETF
    (top 10, pays, secteurs) + date "As of" -> controle de fraicheur.
  - iShares (CSV holdings) : si l'ISIN figure dans ISHARES_PRODUCTID, on utilise
    le fichier holdings complet de l'emetteur (pays/secteurs/top10 EXACTS).
    -> donnees emetteur, plus precises (priorite sur justETF).

FRAICHEUR
---------
Pour chaque ETF, la date de la donnee doit etre POSTERIEURE a MIN_DATE
(01/01/2026 par defaut). Si la donnee est plus ancienne (ou introuvable),
l'ETF est IGNORE : on ne remplace jamais une valeur par une donnee perimee.

TICKER ALTERNATIF / ETF SYNTHETIQUES
------------------------------------
Les ETF synthetiques (swap) ne publient pas leurs positions reelles. On utilise
alors un "proxy" physique (ticker/ISIN alternatif) : voir SYNTHETIC_PROXY +
colonne "ticker alternatif" du fichier.

Dependances :  pip install requests beautifulsoup4 openpyxl google-auth google-api-python-client
"""

import os
import re
import csv
import io
import json
import datetime
import unicodedata

import requests

try:
    from bs4 import BeautifulSoup
    BS4_OK = True
except Exception:
    BS4_OK = False

from openpyxl import load_workbook

try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    GOOGLE_OK = True
except Exception:
    GOOGLE_OK = False

# ==============================================================================
# CONFIGURATION
# ==============================================================================
SHEET_ID = "15w4s6chCytFKmPSpGXeYQ9fiJEVD_T5U9671Q0chn_Q"
SERVICE_ACCOUNT_FILE = "service_account.json"     # ou variable d'env GOOGLE_SERVICE_ACCOUNT

# Fichier Excel a mettre a jour (cherche a cote du script par defaut)
HERE = os.path.dirname(os.path.abspath(__file__))
XLSX_FILE = os.environ.get("ETF_XLSX", os.path.join(HERE, "repartition_ETFs.xlsx"))
# Dashboard d'ou on lit ETF_META (registre des ETF). A cote du script par defaut.
DASHBOARD_HTML = os.environ.get("DASHBOARD_HTML", os.path.join(HERE, "dashboard_portefeuille.html"))

# Fraicheur minimale : on ignore toute donnee datee <= a cette date (01/01/2026).
MIN_DATE = datetime.date(2026, 1, 1)

# ETF synthetiques -> ISIN du proxy physique a utiliser pour la composition.
#   cle = ISIN du fonds detenu ; valeur = ISIN du proxy a interroger.
SYNTHETIC_PROXY = {
    "LU2009202107": "IE00BMG6Z448",   # Amundi MSCI EM ex China (synth) -> iShares MSCI ex China
    "IE00BMTX1Y45": "IE00B3XXRP09",   # iShares S&P500 Swap (synth)     -> Vanguard S&P 500 (VUSA)
}

# Enrichissement emetteur iShares : ISIN -> productId iShares (page produit).
# Quand renseigne, on telecharge le CSV holdings complet (pays/secteurs/top10 exacts).
# Pour ajouter un fonds : ouvrir sa fiche sur ishares.com et relever le numero
# dans l'URL  .../products/<productId>/...
ISHARES_PRODUCTID = {
    "IE00B1XNHC34": "251382",   # iShares Global Clean Energy (IQQH)              [verifie 18/06/2026]
    "IE000U58J0M1": "326325",   # iShares Global Clean Energy Transition (Q8Y0)   [verifie 18/06/2026]
    "IE00BMG6Z448": "315592",   # iShares MSCI EM ex-China (MTPI) = proxy d'EMXC  [verifie 18/06/2026]
}

ISHARES_BASE = "https://www.ishares.com/uk/individual/en/products/{pid}/fund/1506575576011.ajax?fileType=csv&fileName={pid}_holdings&dataType=fund"

UA = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                    "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"}

# Correspondance des secteurs (taxonomies GICS / justETF) -> colonnes du fichier.
SECTOR_MAP = {
    # GICS (iShares)
    "information technology": "Technology",
    "technology": "Technology",
    "industrials": "Industry",
    "industry": "Industry",
    "materials": "Materials",
    "basic materials": "Materials",
    "energy": "Energy",
    "communication": "Telecommunication",
    "communication services": "Telecommunication",
    "telecommunication": "Telecommunication",
    "telecommunications": "Telecommunication",
    "health care": "Health",
    "healthcare": "Health",
    "health": "Health",
    "financials": "Banking",
    "financial services": "Banking",
    "banking": "Banking",
    "consumer discretionary": "Cyclical",
    "consumer cyclical": "Cyclical",
    "consumer cyclicals": "Cyclical",
    "cyclical": "Cyclical",
    "consumer staples": "Non-Cyclical",
    "consumer non-cyclical": "Non-Cyclical",
    "consumer non-cyclicals": "Non-Cyclical",
    "non-cyclical": "Non-Cyclical",
    "real estate": "Real Estate",
    "utilities": "Utilities",
}
CASH_SECTORS = {"cash and/or derivatives", "cash", "money market", "other", "others", ""}

# Normalisation des noms de PAYS (source -> nom de colonne du fichier).
COUNTRY_MAP = {
    "korea (south)": "South Korea",
    "korea, republic of": "South Korea",
    "republic of korea": "South Korea",
    "korea": "South Korea",
    "united states of america": "United States",
    "usa": "United States",
    "russian federation": "Russia",
    "taiwan (province of china)": "Taiwan",
    "taiwan, province of china": "Taiwan",
    "hong kong sar": "Hong Kong",
    "viet nam": "Vietnam",
    "czechia": "Czech Republic",
    "other": "others",
    "others": "others",
}

# Colonnes "metadonnees" (jamais traitees comme des donnees pays/secteur).
META_COLS = {"nom", "isin", "ticker", "ticker alternatif"}

# Format pourcentage uniforme applique a toutes les cellules de donnees.
PCT_FMT = "0.00%"


def log(msg=""):
    print(msg, flush=True)


# ==============================================================================
# 1) GOOGLE SHEETS : lecture du portefeuille (book2_all_rows + book2_extra_rows)
# ==============================================================================
def _get_creds():
    SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
    sa_env = os.environ.get("GOOGLE_SERVICE_ACCOUNT")
    if sa_env:
        return service_account.Credentials.from_service_account_info(json.loads(sa_env), scopes=SCOPES)
    return service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)


def lire_portefeuille():
    """Renvoie la liste des positions [{isin,ticker,name,typeInv,broker,...}, ...]."""
    if not GOOGLE_OK:
        log("  [Data] librairies Google absentes -> portefeuille indisponible.")
        return None
    try:
        svc = build("sheets", "v4", credentials=_get_creds()).spreadsheets()
        res = svc.values().get(spreadsheetId=SHEET_ID, range="Data!A:Z").execute()
        data = {}
        for r in (res.get("values", []) or []):
            if r and r[0]:
                data[r[0]] = "".join(r[1:])      # recoller les morceaux (chunks)
        rows = []
        for key in ("book2_all_rows", "book2_extra_rows"):
            raw = data.get(key)
            if raw:
                try:
                    lst = json.loads(raw)
                    if isinstance(lst, list):
                        rows += lst
                except Exception:
                    pass
        return rows or None
    except Exception as e:
        log(f"  [Data] lecture portefeuille impossible : {e}")
        return None


# ==============================================================================
# 2) ETF_META : registre des ETF (ticker -> isin, name) lu depuis le dashboard
# ==============================================================================
def lire_etf_meta():
    """Extrait ETF_META de dashboard_portefeuille.html -> {ticker: {isin, name}}."""
    meta = {}
    try:
        html = open(DASHBOARD_HTML, encoding="utf-8").read()
    except Exception as e:
        log(f"  [ETF_META] lecture {DASHBOARD_HTML} impossible : {e}")
        return meta
    md = re.search(r"ETF_META\s*=\s*\{", html)   # l'ASSIGNATION (pas une simple mention)
    if not md:
        return meta
    i = html.find("{", md.start())
    # extraction equilibree des accolades de l'objet ETF_META
    depth, j, instr, esc = 0, i, False, False
    for j in range(i, min(len(html), i + 400000)):
        c = html[j]
        if instr:
            if esc:
                esc = False
            elif c == "\\":
                esc = True
            elif c == "'":
                instr = False
        else:
            if c == "'":
                instr = True
            elif c == "{":
                depth += 1
            elif c == "}":
                depth -= 1
                if depth == 0:
                    break
    blob = html[i:j + 1]
    # chaque entree : 'TICKER': { ... [isin:'XX'] name:'YY' ... }
    for m in re.finditer(r"['\"]?([A-Za-z0-9.\-]+)['\"]?\s*:\s*\{", blob):
        tk = m.group(1)
        seg = blob[m.end():m.end() + 800]
        isin = re.search(r"isin\s*:\s*'([^']*)'", seg)
        name = re.search(r"name\s*:\s*'([^']*)'", seg)
        meta[tk] = {"isin": (isin.group(1).upper() if isin else ""),
                    "name": (name.group(1) if name else tk)}
    return meta


# ==============================================================================
# 3) Construction de la liste DYNAMIQUE des ETF detenus
# ==============================================================================
def etfs_du_portefeuille():
    """[{isin, ticker, name}] des ETF detenus (dedup par ISIN). None si indispo."""
    hold = lire_portefeuille()
    if not hold:
        return None
    meta = lire_etf_meta()
    meta_isins = {v["isin"] for v in meta.values()}
    seen = {}
    for r in hold:
        if not isinstance(r, dict):
            continue
        tk = (r.get("ticker") or "").strip()
        isin = (r.get("isin") or "").strip().upper()
        ty = (r.get("typeInv") or "").strip().lower()
        is_etf = (tk in meta) or (isin in meta_isins) or (ty == "etf")
        if not is_etf or not isin:
            continue
        if isin not in seen:
            seen[isin] = {"isin": isin, "ticker": tk or isin,
                          "name": r.get("name") or (meta.get(tk, {}) or {}).get("name") or tk}
    return list(seen.values())


# ==============================================================================
# 4) PARSERS de composition
# ==============================================================================
def _num(s):
    try:
        return float(str(s).replace(",", "").strip())
    except Exception:
        return None


def parse_ishares_csv(text):
    """CSV holdings iShares -> dict {date, countries{}, sectors{}, top10[(name,w)]} (poids en fraction)."""
    lines = text.splitlines()
    # date : 1re ligne  Fund Holdings as of,"DD/Mon/YYYY"
    d = None
    m = re.search(r'as of[",\s]+(\d{1,2}[/ ][A-Za-z]{3}[/ ]\d{4})', text[:200])
    if m:
        for fmt in ("%d/%b/%Y", "%d %b %Y"):
            try:
                d = datetime.datetime.strptime(m.group(1).replace(" ", "/"), fmt).date()
                break
            except Exception:
                pass
    # localiser l'entete du tableau
    hdr_i = next((k for k, ln in enumerate(lines) if ln.startswith("Ticker,Name,")), None)
    if hdr_i is None:
        return None
    reader = csv.DictReader(lines[hdr_i:])
    countries, sectors, holdings = {}, {}, []
    for row in reader:
        ac = (row.get("Asset Class") or "").strip().lower()
        w = _num(row.get("Weight (%)"))
        if w is None:
            continue
        wf = w / 100.0
        name = (row.get("Name") or "").strip()
        sec = (row.get("Sector") or "").strip()
        loc = (row.get("Location") or "").strip()
        if ac == "equity":
            holdings.append((name, wf))
            if loc:
                countries[loc] = countries.get(loc, 0.0) + wf
            mapped = SECTOR_MAP.get(_norm(sec))
            if mapped:
                sectors[mapped] = sectors.get(mapped, 0.0) + wf
            else:
                sectors["Others"] = sectors.get("Others", 0.0) + wf
    holdings.sort(key=lambda x: x[1], reverse=True)
    return {"date": d, "countries": countries, "sectors": sectors, "top10": holdings[:10]}


def _norm(s):
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().lower().strip()
    return s


def parse_justetf(html):
    """Page profil justETF -> dict {date, countries{}, sectors{}, top10[(name,w)]} (poids en fraction)."""
    date = None
    m = re.search(r"As of\s+(\d{2})/(\d{2})/(\d{4})", html)
    if m:
        date = datetime.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))

    countries, sectors, top10 = {}, {}, []
    if BS4_OK:
        soup = BeautifulSoup(html, "html.parser")

        def rows_after(title):
            h = soup.find(lambda t: t.name in ("h2", "h3", "h4") and t.get_text(strip=True) == title)
            if not h:
                return []
            tbl = h.find_next("table")
            out = []
            if tbl:
                for tr in tbl.find_all("tr"):
                    tds = [td.get_text(strip=True) for td in tr.find_all("td")]
                    if len(tds) >= 2:
                        out.append((tds[0], tds[-1]))
            return out

        for name, pct in rows_after("Countries"):
            v = _pct(pct)
            if v is not None:
                key = "others" if name.lower() in ("other", "others") else name
                countries[key] = countries.get(key, 0.0) + v
        for name, pct in rows_after("Sectors"):
            v = _pct(pct)
            if v is None:
                continue
            mapped = SECTOR_MAP.get(_norm(name), "Others")
            sectors[mapped] = sectors.get(mapped, 0.0) + v
        # Top 10 : table de la section "Top 10 Holdings"
        h = soup.find(lambda t: t.name in ("h2", "h3", "h4") and "Top 10 Holdings" in t.get_text())
        if h:
            tbl = h.find_next("table")
            if tbl:
                for tr in tbl.find_all("tr"):
                    tds = [td.get_text(strip=True) for td in tr.find_all("td")]
                    if len(tds) >= 2 and _pct(tds[-1]) is not None and "top 10" not in tds[0].lower():
                        top10.append((tds[0], _pct(tds[-1])))
    return {"date": date, "countries": countries, "sectors": sectors, "top10": top10[:10]}


def _pct(s):
    m = re.search(r"(-?\d+(?:[.,]\d+)?)\s*%", str(s))
    return float(m.group(1).replace(",", ".")) / 100.0 if m else None


# ==============================================================================
# 5) Recuperation de la composition d'un ETF (iShares -> sinon justETF)
# ==============================================================================
def build_alt_map(wb):
    """Lit la colonne 'ticker alternatif' du fichier -> {isin_fonds: isin_proxy}.

    Le 'ticker alternatif' est un TICKER ; on le resout en ISIN grace a la table
    ticker->ISIN construite a partir des onglets du fichier."""
    tk2isin = {}
    for sh in ("by country", "by sector"):
        if sh not in wb.sheetnames:
            continue
        ws = wb[sh]
        ctk, cisin = col_index(ws, "ticker"), col_index(ws, "ISIN")
        if not (ctk and cisin):
            continue
        for r in range(2, ws.max_row + 1):
            tk = str(ws.cell(r, ctk).value or "").strip()
            isin = str(ws.cell(r, cisin).value or "").strip().upper()
            if tk and isin:
                tk2isin[tk] = isin
    alt = {}
    for sh in ("by country", "by sector"):
        if sh not in wb.sheetnames:
            continue
        ws = wb[sh]
        cisin, calt = col_index(ws, "ISIN"), col_index(ws, "ticker alternatif")
        if not (cisin and calt):
            continue
        for r in range(2, ws.max_row + 1):
            isin = str(ws.cell(r, cisin).value or "").strip().upper()
            altk = str(ws.cell(r, calt).value or "").strip()
            if isin and altk and tk2isin.get(altk):
                alt[isin] = tk2isin[altk]
    return alt


def fetch_compo(isin, proxy=None):
    """Renvoie (compo, source) ou (None, raison)."""
    proxy = proxy or SYNTHETIC_PROXY
    data_isin = proxy.get(isin, isin)

    # a) iShares (emetteur) si productId connu
    pid = ISHARES_PRODUCTID.get(data_isin)
    if pid:
        try:
            url = ISHARES_BASE.format(pid=pid)
            r = requests.get(url, headers=UA, timeout=30)
            if r.ok and "Weight" in r.text:
                compo = parse_ishares_csv(r.text)
                if compo and (compo["countries"] or compo["top10"]):
                    return compo, f"iShares (CSV, productId {pid})"
        except Exception as e:
            log(f"    [iShares {data_isin}] {e}")

    # b) justETF (universel, par ISIN)
    try:
        url = f"https://www.justetf.com/en/etf-profile.html?isin={data_isin}"
        r = requests.get(url, headers=UA, timeout=30)
        if r.ok:
            compo = parse_justetf(r.text)
            if compo and (compo["countries"] or compo["sectors"] or compo["top10"]):
                return compo, "justETF"
    except Exception as e:
        log(f"    [justETF {data_isin}] {e}")
    return None, "aucune source exploitable"


def date_ok(compo):
    d = compo.get("date")
    return bool(d) and d > MIN_DATE


# ==============================================================================
# 6) Mise a jour du fichier Excel
# ==============================================================================
def col_index(ws, header, row=1):
    for c in range(1, ws.max_column + 1):
        if (ws.cell(row, c).value or "") == header:
            return c
    return None


def ensure_col(ws, header, row=1):
    c = col_index(ws, header, row)
    if c:
        return c
    c = ws.max_column + 1
    ws.cell(row, c).value = header
    return c


def find_etf_row(ws, isin, ticker):
    """Trouve la ligne de l'ETF (par ISIN puis ticker). Cree si absente."""
    cisin, ctk = col_index(ws, "ISIN"), col_index(ws, "ticker")
    for r in range(2, ws.max_row + 1):
        if cisin and str(ws.cell(r, cisin).value or "").upper() == isin:
            return r
        if ctk and str(ws.cell(r, ctk).value or "").strip() == ticker:
            return r
    r = ws.max_row + 1
    return r  # nouvelle ligne (remplie par l'appelant)


def _data_cols(ws):
    """Indices des colonnes de DONNEES (en-tete non vide et hors metadonnees)."""
    out = []
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(1, c).value or "").strip()
        if h and h.lower() not in META_COLS:
            out.append(c)
    return out


def _find_col_ci(ws, name):
    """Trouve une colonne par en-tete (insensible a la casse)."""
    nl = str(name).strip().lower()
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value or "").strip().lower() == nl:
            return c
    return None


def enforce_pct_format(ws):
    """Applique le format pourcentage uniforme a toutes les cellules de donnees."""
    cols = _data_cols(ws)
    for c in cols:
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, c).value is not None:
                ws.cell(r, c).number_format = PCT_FMT


def adjust_to_100(weights, others_key):
    """Force le total a 100% en ajustant la colonne 'others' / 'Others'
    (comble notamment le cash exclu des fichiers holdings)."""
    if not weights:
        return weights
    total = sum(v for v in weights.values() if v)
    resid = round(1.0 - total, 6)
    new_other = round(weights.get(others_key, 0.0) + resid, 6)
    if new_other < 0:               # somme > 100% (rare) : on ne descend pas sous 0
        new_other = 0.0
    if abs(resid) > 1e-6 or others_key in weights:
        weights[others_key] = new_other
    return weights


def update_breakdown_sheet(ws, etf, weights, kind):
    """Onglet 'by country' / 'by sector' : MAJ d'une ligne ETF.
    - structure preservee (colonnes meta detectees par en-tete) ;
    - noms de pays normalises (evite les doublons type 'Korea (South)') ;
    - format pourcentage uniforme."""
    cnom, cisin, ctk = col_index(ws, "Nom"), col_index(ws, "ISIN"), col_index(ws, "ticker")
    r = find_etf_row(ws, etf["isin"], etf["ticker"])
    if cnom: ws.cell(r, cnom).value = etf["name"]
    if cisin: ws.cell(r, cisin).value = etf["isin"]
    if ctk: ws.cell(r, ctk).value = etf["ticker"]
    # vider uniquement les colonnes de DONNEES de cette ligne
    for c in _data_cols(ws):
        ws.cell(r, c).value = None
    added = []
    for label, w in weights.items():
        if w is None or w <= 0:
            continue
        name = label
        if kind == "country":
            name = COUNTRY_MAP.get(str(label).strip().lower(), label)
        c = _find_col_ci(ws, name)
        if not c:
            c = ensure_col(ws, name)
            added.append(name)
        cell = ws.cell(r, c)
        cell.value = round(w, 4)
        cell.number_format = PCT_FMT
    return added


def update_top10_sheet(ws, etf, top10):
    """Onglet 'TOP 10 position' (ETF en colonnes : ligne1 TICKER, l2 ALT, l3 ISIN)."""
    # trouver/creer la colonne de l'ETF (par ISIN ligne 3, sinon ticker ligne 1)
    col = None
    for c in range(2, ws.max_column + 1):
        if str(ws.cell(3, c).value or "").upper() == etf["isin"] or \
           str(ws.cell(1, c).value or "").strip() == etf["ticker"]:
            col = c
            break
    if not col:
        col = ws.max_column + 1
        ws.cell(1, col).value = etf["ticker"]
        ws.cell(3, col).value = etf["isin"]
    # vider l'ancienne colonne (a partir de la ligne 4)
    for r in range(4, ws.max_row + 1):
        ws.cell(r, col).value = None
    # index des noms de positions (colonne 1, a partir de la ligne 4)
    name_rows = {}
    for r in range(4, ws.max_row + 1):
        nm = ws.cell(r, 1).value
        if nm:
            name_rows[str(nm).strip().lower()] = r
    for name, w in top10:
        key = str(name).strip().lower()
        r = name_rows.get(key)
        if not r:
            r = ws.max_row + 1
            ws.cell(r, 1).value = name
            name_rows[key] = r
        cell = ws.cell(r, col)
        cell.value = round(w, 4)
        cell.number_format = PCT_FMT


# ==============================================================================
# MAIN
# ==============================================================================
def main():
    log("=" * 70)
    log(" Mise a jour des compositions ETF  ->  " + os.path.basename(XLSX_FILE))
    log("=" * 70)

    etfs = etfs_du_portefeuille()
    if etfs is None:
        log("\n  Portefeuille indisponible (Google Sheets). Arret.")
        return
    log(f"\n  {len(etfs)} ETF detectes dans le portefeuille (liste DYNAMIQUE).")

    if not os.path.exists(XLSX_FILE):
        log(f"  ERREUR : fichier introuvable : {XLSX_FILE}")
        return
    wb = load_workbook(XLSX_FILE)
    ws_country = wb["by country"]
    ws_sector = wb["by sector"]
    ws_top10 = wb["TOP 10 position"]

    # proxys "ticker alternatif" (fichier) + synthetiques connus
    proxy = dict(SYNTHETIC_PROXY)
    proxy.update(build_alt_map(wb))

    maj, ignores, new_cols = 0, [], set()
    for etf in sorted(etfs, key=lambda e: e["ticker"]):
        compo, source = fetch_compo(etf["isin"], proxy)
        if not compo:
            log(f"  KO  {etf['ticker']:10s} ({etf['isin']}) -> {source}")
            ignores.append(etf["ticker"]); continue
        if not date_ok(compo):
            log(f"  --  {etf['ticker']:10s} ({etf['isin']}) -> donnee perimee/sans date "
                f"({compo.get('date')}) : IGNORE")
            ignores.append(etf["ticker"]); continue

        # total force a 100% (cash exclu -> bascule dans others/Others)
        compo["countries"] = adjust_to_100(compo["countries"], "others")
        compo["sectors"] = adjust_to_100(compo["sectors"], "Others")
        ac = update_breakdown_sheet(ws_country, etf, compo["countries"], "country")
        asx = update_breakdown_sheet(ws_sector, etf, compo["sectors"], "sector")
        update_top10_sheet(ws_top10, etf, compo["top10"])
        new_cols.update(ac); new_cols.update(asx)
        maj += 1
        log(f"  OK  {etf['ticker']:10s} ({etf['isin']}) -> {source} | as of {compo['date']} | "
            f"{len(compo['countries'])} pays, {len(compo['sectors'])} secteurs, {len(compo['top10'])} top")

    # format pourcentage uniforme sur tout le fichier (coherence d'affichage)
    enforce_pct_format(ws_country)
    enforce_pct_format(ws_sector)

    # Sauvegarde robuste : si le fichier est verrouille (ouvert dans Excel /
    # verrou OneDrive), on ecrit dans une copie "_maj" au lieu de planter.
    out_file = XLSX_FILE
    try:
        wb.save(XLSX_FILE)
    except PermissionError:
        base, ext = os.path.splitext(XLSX_FILE)
        out_file = base + "_maj" + ext
        try:
            wb.save(out_file)
            log("\n  /!\\ Fichier d'origine verrouille (Excel ouvert / OneDrive).")
            log(f"      -> resultat ecrit dans : {out_file}")
            log("      Ferme repartition_ETFs.xlsx puis remplace-le par cette copie,")
            log("      ou ferme Excel et relance pour ecrire directement dans l'original.")
        except PermissionError:
            log("\n  ERREUR : impossible d'ecrire (dossier verrouille). "
                "Ferme Excel / mets OneDrive en pause puis relance.")
            return

    log("\n" + "-" * 70)
    log(f"  Termine : {maj} ETF mis a jour, {len(ignores)} ignores.")
    if new_cols:
        log(f"  Nouvelles colonnes pays/secteur ajoutees : {', '.join(sorted(new_cols))}")
    if ignores:
        log(f"  Ignores (source/fraicheur) : {', '.join(ignores)}")
    log(f"  Fichier sauvegarde : {out_file}")


if __name__ == "__main__":
    main()
# fin
